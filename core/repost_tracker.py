"""
core/repost_tracker.py — Excel-based deduplication tracker.

Stores posted shortcodes in data/tracker.xlsx (sheet: "Repost Log").
Replaces the legacy data/reposted_ids.txt flat-file approach.

Columns: shortcode | post_type | posted_at | source_url | category

On first run, automatically migrates any existing reposted_ids.txt data
into the Excel workbook (deduplicated), then the .txt file can be removed.

Usage:
    from core.repost_tracker import is_reposted, mark_reposted

    if not is_reposted("DU3i6qPDTOH"):
        # process and post...
        mark_reposted("DU3i6qPDTOH", post_type="image", source_url="...", category="dance")
"""

from datetime import datetime, timezone
from pathlib import Path
from typing import List

# ── File / sheet constants ──────────────────────────────────────────────────
TRACKER_FILE   = Path("data/tracker.xlsx")
SHEET_LOG      = "Repost Log"
SHEET_STATE    = "State"

_LOG_HEADERS   = ["shortcode", "post_type", "posted_at", "source_url", "category"]
_STATE_HEADERS = ["key", "value"]


# ── Internal helpers ────────────────────────────────────────────────────────

def _ensure_workbook() -> None:
    """
    Create tracker.xlsx with required sheets if it doesn't exist yet,
    then trigger one-time migration from legacy .txt files.
    """
    import openpyxl

    if TRACKER_FILE.exists():
        return

    TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 1: Repost Log
    ws_log = wb.active
    ws_log.title = SHEET_LOG
    ws_log.append(_LOG_HEADERS)

    # Sheet 2: State (key/value — stores last_post_type etc.)
    ws_state = wb.create_sheet(SHEET_STATE)
    ws_state.append(_STATE_HEADERS)
    # Default: pretend last was a reel → next run will post an image
    ws_state.append(["last_post_type", "reel"])

    wb.save(TRACKER_FILE)

    # One-time migration from legacy .txt files
    _migrate_from_txt()


def _migrate_from_txt() -> None:
    """
    Reads legacy data/reposted_ids.txt (if it exists), deduplicates
    the shortcodes, and inserts them into the Repost Log sheet.
    Also migrates last_post_type.txt into the State sheet.
    Called once when tracker.xlsx is first created.
    """
    import openpyxl

    # ── Migrate reposted_ids.txt ───────────────────────────────────────────
    old_log = Path("data/reposted_ids.txt")
    if old_log.exists():
        raw = [l.strip() for l in old_log.read_text(encoding="utf-8").splitlines()]
        # Deduplicate while preserving order
        seen: set = set()
        unique_ids: List[str] = []
        for sc in raw:
            if sc and sc not in seen:
                seen.add(sc)
                unique_ids.append(sc)

        if unique_ids:
            wb = openpyxl.load_workbook(TRACKER_FILE)
            ws = wb[SHEET_LOG]
            for shortcode in unique_ids:
                ws.append([shortcode, "unknown", "migrated", "", ""])
            wb.save(TRACKER_FILE)

    # ── Migrate last_post_type.txt ─────────────────────────────────────────
    old_state = Path("data/last_post_type.txt")
    if old_state.exists():
        last_type = old_state.read_text(encoding="utf-8").strip().lower()
        if last_type in ("image", "reel"):
            wb = openpyxl.load_workbook(TRACKER_FILE)
            ws = wb[SHEET_STATE]
            for row in ws.iter_rows(min_row=2):
                if row[0].value == "last_post_type":
                    row[1].value = last_type
                    break
            wb.save(TRACKER_FILE)


# ── Public API ──────────────────────────────────────────────────────────────

def is_reposted(shortcode: str) -> bool:
    """Return True if this shortcode has already been posted."""
    _ensure_workbook()
    import openpyxl

    shortcode = shortcode.strip()
    wb = openpyxl.load_workbook(TRACKER_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_LOG]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and str(row[0]).strip() == shortcode:
            wb.close()
            return True
    wb.close()
    return False


def mark_reposted(
    shortcode:  str,
    post_type:  str = "unknown",
    source_url: str = "",
    category:   str = "",
) -> None:
    """
    Append a new row to the Repost Log sheet.

    Args:
        shortcode:  Instagram shortcode (e.g. 'DU3i6qPDTOH').
        post_type:  'image' or 'reel'.
        source_url: Full Instagram URL (for reference).
        category:   Caption category used (dance/humor/lifestyle/trending/motivation/general).
    """
    _ensure_workbook()
    import openpyxl

    posted_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    wb = openpyxl.load_workbook(TRACKER_FILE)
    ws = wb[SHEET_LOG]
    ws.append([shortcode.strip(), post_type, posted_at, source_url, category])
    wb.save(TRACKER_FILE)


def all_reposted() -> List[str]:
    """Return all tracked shortcodes (for debugging / inspection)."""
    _ensure_workbook()
    import openpyxl

    wb = openpyxl.load_workbook(TRACKER_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_LOG]
    result = [
        str(row[0]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    ]
    wb.close()
    return result


def get_last_posted_at(shortcode: str) -> datetime:
    """
    Return the posted_at datetime for a shortcode.

    Used by the Priority 3 safeguard to sort all sheet URLs by age so the
    system re-posts the one that was posted the longest time ago (stalest),
    not the most recently posted content.

    Returns:
        datetime — actual posted_at if available.
        datetime.min — for 'migrated' entries (treat as very old / always eligible).
        datetime.max — for shortcodes not in the tracker (should not happen in P3,
                       but means 'unknown age, prefer last').
    """
    _ensure_workbook()
    import openpyxl

    shortcode = shortcode.strip()
    wb = openpyxl.load_workbook(TRACKER_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_LOG]

    oldest_dt = None   # if a shortcode appears multiple times, use the earliest row

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or str(row[0]).strip() != shortcode:
            continue

        raw = str(row[2]).strip() if row[2] else ""

        if raw == "migrated" or not raw:
            # Treat migrated entries as extremely old so they're picked first
            dt = datetime.min.replace(tzinfo=timezone.utc)
        else:
            try:
                dt = datetime.strptime(raw, "%Y-%m-%d %H:%M:%S UTC").replace(tzinfo=timezone.utc)
            except ValueError:
                dt = datetime.min.replace(tzinfo=timezone.utc)

        if oldest_dt is None or dt < oldest_dt:
            oldest_dt = dt

    wb.close()

    if oldest_dt is None:
        # Shortcode not in tracker — treat as newest (least preferred for repeat)
        return datetime.max.replace(tzinfo=timezone.utc)

    return oldest_dt

