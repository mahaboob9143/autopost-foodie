"""
core/post_state.py — Tracks the last post type to enforce alternating pattern.

Reads/writes the 'State' sheet in data/tracker.xlsx.
Replaces the legacy data/last_post_type.txt flat file.

Pattern: image → reel → image → reel → ...
"""

from pathlib import Path

from core.repost_tracker import TRACKER_FILE, SHEET_STATE, _ensure_workbook

_KEY = "last_post_type"


def get_last_post_type() -> str:
    """Return 'image' or 'reel' — whichever was posted last."""
    _ensure_workbook()
    import openpyxl

    wb = openpyxl.load_workbook(TRACKER_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_STATE]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == _KEY:
            wb.close()
            return str(row[1]).strip().lower() if row[1] else "reel"
    wb.close()
    return "reel"  # safe default → next run will post an image


def get_next_post_type() -> str:
    """Return which type should be posted next (opposite of last)."""
    last = get_last_post_type()
    return "reel" if last == "image" else "image"


def save_post_type(post_type: str) -> None:
    """
    Update the State sheet with the type that was just published.
    Creates the row if it doesn't exist yet.
    """
    _ensure_workbook()
    import openpyxl

    post_type = post_type.strip().lower()
    wb = openpyxl.load_workbook(TRACKER_FILE)
    ws = wb[SHEET_STATE]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == _KEY:
            row[1].value = post_type
            wb.save(TRACKER_FILE)
            return

    # Key not found — append it (shouldn't happen after _ensure_workbook, but be safe)
    ws.append([_KEY, post_type])
    wb.save(TRACKER_FILE)
